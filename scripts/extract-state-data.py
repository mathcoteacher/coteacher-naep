#!/usr/bin/env python3
"""
Extract school-level 8th grade math proficiency data for each state,
join with NCES coordinates, and output intermediate JSON files.

Reads:
  - NCES EDGE geocode data (lat/lng per school)
  - NCES CCD school directory (state ID → NCES ID crosswalk)
  - State achievement xlsx/csv files

Outputs:
  - scripts/intermediate/{STATE}.json with lat/lng (not yet projected)
"""

import csv
import json
import os
import sys
from pathlib import Path

# Try to import openpyxl for xlsx files
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl required. Install with: pip3 install openpyxl")
    sys.exit(1)

# Paths
EDGE_FILE = "/Volumes/SignatureMi/ohio_education_data/data/raw/nces_edge/edge_geocode_2425.csv"
CCD_FILE = "/Volumes/SignatureMi/ohio_education_data/data/raw/nces_ccd/ccd_sch_029_2425_w_0a_051425.csv"
ACHIEVEMENT_DIR = "/Volumes/SignatureMi/ohio_education_data/data/raw/state_achievement"
NAEP_FILE = os.path.join(os.path.dirname(__file__), "..", "public", "naep.json")
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "intermediate")

# FIPS state codes for name lookup
FIPS_TO_STATE = {
    "01": "AL", "02": "AK", "04": "AZ", "05": "AR", "06": "CA",
    "08": "CO", "09": "CT", "10": "DE", "11": "DC", "12": "FL",
    "13": "GA", "15": "HI", "16": "ID", "17": "IL", "18": "IN",
    "19": "IA", "20": "KS", "21": "KY", "22": "LA", "23": "ME",
    "24": "MD", "25": "MA", "26": "MI", "27": "MN", "28": "MS",
    "29": "MO", "30": "MT", "31": "NE", "32": "NV", "33": "NH",
    "34": "NJ", "35": "NM", "36": "NY", "37": "NC", "38": "ND",
    "39": "OH", "40": "OK", "41": "OR", "42": "PA", "44": "RI",
    "45": "SC", "46": "SD", "47": "TN", "48": "TX", "49": "UT",
    "50": "VT", "51": "VA", "53": "WA", "54": "WV", "55": "WI",
    "56": "WY"
}

STATE_NAMES = {
    "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
    "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
    "DC": "District of Columbia", "FL": "Florida", "GA": "Georgia", "HI": "Hawaii",
    "ID": "Idaho", "IL": "Illinois", "IN": "Indiana", "IA": "Iowa",
    "KS": "Kansas", "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine",
    "MD": "Maryland", "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota",
    "MS": "Mississippi", "MO": "Missouri", "MT": "Montana", "NE": "Nebraska",
    "NV": "Nevada", "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico",
    "NY": "New York", "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio",
    "OK": "Oklahoma", "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island",
    "SC": "South Carolina", "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas",
    "UT": "Utah", "VT": "Vermont", "VA": "Virginia", "WA": "Washington",
    "WV": "West Virginia", "WI": "Wisconsin", "WY": "Wyoming"
}


def load_edge_data():
    """Load NCES EDGE geocode data: NCESSCH → {lat, lon, name, city}"""
    print("Loading EDGE geocode data...")
    edge = {}
    with open(EDGE_FILE) as f:
        reader = csv.DictReader(f, delimiter='|')
        for row in reader:
            ncessch = row['NCESSCH'].strip()
            try:
                lat = float(row['LAT'])
                lon = float(row['LON'])
            except (ValueError, KeyError):
                continue
            edge[ncessch] = {
                'lat': lat,
                'lon': lon,
                'name': row['NAME'].strip(),
                'city': row['CITY'].strip(),
                'state': row['STATE'].strip()
            }
    print(f"  Loaded {len(edge)} school coordinates")
    return edge


def load_ccd_crosswalk():
    """Load CCD data for state school ID → NCESSCH crosswalk"""
    print("Loading CCD crosswalk...")
    # Build two lookups:
    # 1. ST_SCHID → NCESSCH (for states that use state IDs)
    # 2. NCESSCH → {district, city} (for metadata)
    st_schid_to_ncessch = {}
    ncessch_meta = {}
    with open(CCD_FILE) as f:
        reader = csv.DictReader(f)
        for row in reader:
            ncessch = row['NCESSCH'].strip()
            st_schid = row['ST_SCHID'].strip()
            st_schid_to_ncessch[st_schid] = ncessch
            ncessch_meta[ncessch] = {
                'district': row['LEA_NAME'].strip(),
                'city': row['LCITY'].strip(),
                'state': row['ST'].strip(),
                'name': row['SCH_NAME'].strip(),
                'leaid': row['LEAID'].strip()
            }
    print(f"  Loaded {len(st_schid_to_ncessch)} state ID mappings")
    return st_schid_to_ncessch, ncessch_meta


def load_naep_data():
    """Load NAEP data for state-level stats"""
    with open(NAEP_FILE) as f:
        return json.load(f)


def parse_proficiency(val):
    """Parse a proficiency value, returning float or None"""
    if val is None:
        return None
    s = str(val).strip()
    if s in ('NC', 'N/A', '', '***', '--', '*', 'NA', 'S', 'n/a', '<10'):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# === State-specific extractors ===

def extract_ohio(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract Ohio school-level 8th grade math proficiency"""
    xlsx_path = os.path.join(ACHIEVEMENT_DIR, "OH", "24-25_Achievement_Building.xlsx")
    print(f"  Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Report_Only_Indicators']

    schools = []
    header = None
    math_8_col = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            header = list(row)
            # Find "8th Grade Math 2024-2025 Percent Proficient or above - Building"
            for j, h in enumerate(header):
                if h and '8th Grade Math' in str(h) and '2024-2025' in str(h) and 'Building' in str(h):
                    math_8_col = j
                    break
            if math_8_col is None:
                print("  ERROR: Could not find 8th grade math column")
                wb.close()
                return []
            continue

        building_irn = str(row[0]).strip() if row[0] else None
        building_name = str(row[1]).strip() if row[1] else None
        district_irn = str(row[2]).strip() if row[2] else None
        district_name = str(row[3]).strip() if row[3] else None

        if not building_irn or not district_irn:
            continue

        proficiency = parse_proficiency(row[math_8_col])
        if proficiency is None:
            continue

        # Join: OH-{district_irn}-{building_irn} → NCESSCH → EDGE
        st_schid = f"OH-{district_irn}-{building_irn}"
        ncessch = st_schid_to_ncessch.get(st_schid)
        if not ncessch:
            continue

        if ncessch not in edge:
            continue

        coord = edge[ncessch]
        meta = ncessch_meta.get(ncessch, {})

        schools.append({
            'name': building_name,
            'lat': coord['lat'],
            'lon': coord['lon'],
            'proficiency': round(proficiency / 100, 4),
            'district': district_name or meta.get('district', ''),
            'city': coord.get('city', '') or meta.get('city', ''),
            'ncessch': ncessch
        })

    wb.close()
    return schools


def extract_pennsylvania(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract Pennsylvania school-level 8th grade math proficiency"""
    xlsx_path = os.path.join(ACHIEVEMENT_DIR, "PA", "pennsylvania_pssa_school_2024.xlsx")
    print(f"  Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    schools = []
    header = None
    header_row_idx = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = list(row)
        if vals and vals[0] == 'Year':
            header = vals
            header_row_idx = i
            continue
        if header is None:
            continue

        year = vals[0]
        aun = str(vals[1]).strip() if vals[1] else None
        school_number = str(vals[2]).strip() if vals[2] else None
        district_name = str(vals[4]).strip() if vals[4] else None
        school_name = str(vals[5]).strip() if vals[5] else None
        subject = str(vals[6]).strip() if vals[6] else None
        group = str(vals[7]).strip() if vals[7] else None
        grade = vals[8]
        pct_prof_above = vals[14]  # "Percent Proficient and above"

        # Filter: Math, All Students, Grade 8
        if subject != 'Math' or group != 'All Students':
            continue
        if str(grade) != '8':
            continue

        proficiency = parse_proficiency(pct_prof_above)
        if proficiency is None:
            continue

        # PA ST_SCHID format: PA-{AUN}-{school_number_no_leading_zeros}
        # CCD uses e.g. PA-112011103-6921, achievement data has 000006921
        sch_num_stripped = str(int(school_number)) if school_number else ''
        st_schid = f"PA-{aun}-{sch_num_stripped}"
        ncessch = st_schid_to_ncessch.get(st_schid)
        if not ncessch:
            continue

        if ncessch not in edge:
            continue

        coord = edge[ncessch]
        meta = ncessch_meta.get(ncessch, {})

        schools.append({
            'name': school_name,
            'lat': coord['lat'],
            'lon': coord['lon'],
            'proficiency': round(proficiency / 100, 4),
            'district': district_name or meta.get('district', ''),
            'city': coord.get('city', '') or meta.get('city', ''),
            'ncessch': ncessch
        })

    wb.close()
    return schools


def extract_indiana(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract Indiana school-level 8th grade math proficiency"""
    xlsx_path = os.path.join(ACHIEVEMENT_DIR, "IN", "ILEARN-2025-School-All.xlsx")
    print(f"  Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Math']

    schools = []
    header = None
    grade8_prof_col = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 5:
            continue
        vals = list(row)

        if i == 5:
            # This is the header row
            header = vals
            # Indiana Math sheet has grade-specific sections
            # Each grade block: Below, Approaching, At, Above, Total Proficient, Total Tested, %
            # Grades 3,4,5,6,7,8 + All Grades = 7 blocks × 7 cols = 49 cols + 4 id cols = 53
            # Grade 8 is the 6th grade block (0-indexed: 5), starting at col 4 + 5*7 = 39
            # Proficiency % is the 7th col of each block (index 6 within block)
            # So grade 8 proficiency % = col 4 + 5*7 + 6 = 45
            grade8_prof_col = 4 + 5 * 7 + 6  # = 45
            continue

        if header is None:
            continue

        corp_id = str(vals[0]).strip() if vals[0] else None
        corp_name = str(vals[1]).strip() if vals[1] else None
        school_id = str(vals[2]).strip() if vals[2] else None
        school_name = str(vals[3]).strip() if vals[3] else None

        if not corp_id or not school_id:
            continue

        proficiency = parse_proficiency(vals[grade8_prof_col] if grade8_prof_col < len(vals) else None)
        if proficiency is None:
            continue

        # IN ST_SCHID format: IN-{corp_id}-{school_id}
        # Need to find the right format
        st_schid = f"IN-{corp_id}-{school_id}"
        ncessch = st_schid_to_ncessch.get(st_schid)
        if not ncessch:
            continue

        if ncessch not in edge:
            continue

        coord = edge[ncessch]
        meta = ncessch_meta.get(ncessch, {})

        schools.append({
            'name': school_name,
            'lat': coord['lat'],
            'lon': coord['lon'],
            'proficiency': round(proficiency, 4),  # Already 0-1 range
            'district': corp_name or meta.get('district', ''),
            'city': coord.get('city', '') or meta.get('city', ''),
            'ncessch': ncessch
        })

    wb.close()
    return schools


def extract_south_carolina(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract South Carolina school-level 8th grade math proficiency"""
    xlsx_path = os.path.join(ACHIEVEMENT_DIR, "SC", "SC_Achievement_2025.xlsx")
    if not os.path.exists(xlsx_path):
        # Try other file names
        candidates = [f for f in os.listdir(os.path.join(ACHIEVEMENT_DIR, "SC")) if f.endswith('.xlsx')]
        if not candidates:
            print("  No SC xlsx files found")
            return []
        xlsx_path = os.path.join(ACHIEVEMENT_DIR, "SC", candidates[0])

    print(f"  Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)

    # Inspect sheets
    schools = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        # Read header
        header = None
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            vals = list(row)
            # Look for header row with school identifiers
            if any(v and 'School' in str(v) for v in vals[:5]) and any(v and ('Math' in str(v) or 'Proficien' in str(v) or 'Grade' in str(v)) for v in vals):
                header = vals
                break
        if header:
            print(f"  Sheet '{sheet_name}' header: {header[:10]}")
            break

    wb.close()
    return schools


def extract_new_jersey(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract New Jersey school-level 8th grade math proficiency"""
    xlsx_path = os.path.join(ACHIEVEMENT_DIR, "NJ")
    # NJ has grade/subject specific files
    candidates = [f for f in os.listdir(xlsx_path) if '8' in f and 'math' in f.lower() and f.endswith('.xlsx')]
    if not candidates:
        candidates = [f for f in os.listdir(xlsx_path) if '8' in f and f.endswith('.xlsx')]
    if not candidates:
        print(f"  No NJ grade 8 math file found. Files: {os.listdir(xlsx_path)}")
        return []

    file_path = os.path.join(xlsx_path, candidates[0])
    print(f"  Reading {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)
    ws = wb.active

    schools = []
    header = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        vals = list(row)
        if header is None:
            # Look for header
            if any(v and 'School' in str(v) for v in vals):
                header = vals
                print(f"  Header: {[str(v)[:30] for v in vals[:10]]}")
            continue

        # Try to find district code, school code, proficiency columns
        # NJ format varies - extract what we can
        pass

    wb.close()
    return schools


def extract_colorado(edge, st_schid_to_ncessch, ncessch_meta):
    """Extract Colorado school-level 8th grade math proficiency"""
    xlsx_dir = os.path.join(ACHIEVEMENT_DIR, "CO")
    candidates = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx') and 'math' in f.lower()]
    if not candidates:
        candidates = [f for f in os.listdir(xlsx_dir) if f.endswith('.xlsx')]
    if not candidates:
        print(f"  No CO xlsx files found")
        return []

    file_path = os.path.join(xlsx_dir, candidates[0])
    print(f"  Reading {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True)

    print(f"  Sheets: {wb.sheetnames}")
    ws = wb.active
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            print(f"  Row {i}: {list(row)[:10]}")

    wb.close()
    return []


def build_aggregates(schools):
    """Build district and city aggregates from school data"""
    from collections import defaultdict

    # District aggregation
    district_data = defaultdict(lambda: {'schools': [], 'lats': [], 'lons': [], 'profs': []})
    for s in schools:
        d = s['district']
        if d:
            district_data[d]['schools'].append(s)
            district_data[d]['lats'].append(s['lat'])
            district_data[d]['lons'].append(s['lon'])
            district_data[d]['profs'].append(s['proficiency'])

    districts = []
    for name, data in district_data.items():
        n = len(data['schools'])
        districts.append({
            'name': name,
            'lat': sum(data['lats']) / n,
            'lon': sum(data['lons']) / n,
            'proficiency': round(sum(data['profs']) / n, 4),
            'schoolCount': n
        })

    # City aggregation
    city_data = defaultdict(lambda: {'schools': [], 'lats': [], 'lons': [], 'profs': []})
    for s in schools:
        c = s['city']
        if c:
            city_data[c]['schools'].append(s)
            city_data[c]['lats'].append(s['lat'])
            city_data[c]['lons'].append(s['lon'])
            city_data[c]['profs'].append(s['proficiency'])

    cities = []
    for name, data in city_data.items():
        n = len(data['schools'])
        cities.append({
            'name': name,
            'lat': sum(data['lats']) / n,
            'lon': sum(data['lons']) / n,
            'proficiency': round(sum(data['profs']) / n, 4),
            'schoolCount': n
        })

    return districts, cities


def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # Load shared data
    edge = load_edge_data()
    st_schid_to_ncessch, ncessch_meta = load_ccd_crosswalk()
    naep = load_naep_data()

    # State extractors
    extractors = {
        'OH': extract_ohio,
        'PA': extract_pennsylvania,
        'IN': extract_indiana,
    }

    # Process each state
    for state_code, extractor in extractors.items():
        print(f"\n=== Processing {state_code} ({STATE_NAMES.get(state_code, '')}) ===")
        schools = extractor(edge, st_schid_to_ncessch, ncessch_meta)
        print(f"  Extracted {len(schools)} schools with 8th grade math data")

        if not schools:
            print(f"  SKIPPING {state_code} (no data)")
            continue

        districts, cities = build_aggregates(schools)
        print(f"  Aggregated into {len(districts)} districts, {len(cities)} cities")

        # Get NAEP data
        naep_state = naep['states'].get(state_code, naep['national']['US'])

        output = {
            'state': state_code,
            'stateName': STATE_NAMES.get(state_code, state_code),
            'naep': {
                'text': naep_state['text'],
                'numerator': naep_state.get('numerator', 0),
                'denominator': naep_state.get('denominator', 0)
            },
            'schools': schools,
            'districts': districts,
            'cities': cities
        }

        out_path = os.path.join(OUTPUT_DIR, f"{state_code}.json")
        with open(out_path, 'w') as f:
            json.dump(output, f, indent=2)
        print(f"  Wrote {out_path}")

    print("\nDone! Run prepare-map-data.js next to project coordinates.")


if __name__ == '__main__':
    main()
