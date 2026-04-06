"""Tier 1: Nebraska curriculum data from nematerialsmatter.org

Downloads the HQIM (High Quality Instructional Materials) Excel spreadsheet
which lists curriculum adoption by district for ELA, Math, and Science.

Source: https://nematerialsmatter.org/wp-content/uploads/2025/04/2024-25-HQIM-Use-Nebraska-Districts-42225.xlsx

Usage:
    python scripts/curriculum/tier1_nebraska.py
"""

import os
import sqlite3
import ssl
import sys
import urllib.request
from datetime import date

sys.path.insert(0, os.path.dirname(__file__))
from normalize import load_normalization_map, normalize_curriculum_name

DB_PATH = os.path.join(os.path.dirname(__file__), "curriculum.db")
XLSX_URL = "https://nematerialsmatter.org/wp-content/uploads/2025/04/2024-25-HQIM-Use-Nebraska-Districts-42225.xlsx"
XLSX_PATH = os.path.join(os.path.dirname(__file__), "seed_data", "ne_hqim_2024_25.xlsx")
SOURCE_URL = "https://nematerialsmatter.org/im-map/"
SOURCE_TIER = 1
TODAY = date.today().isoformat()


def download_xlsx():
    """Download the Excel file if not already cached."""
    if os.path.exists(XLSX_PATH):
        print(f"Using cached file: {XLSX_PATH}")
        return

    print(f"Downloading {XLSX_URL}...")
    ctx = ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(XLSX_URL, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, context=ctx) as resp:
        data = resp.read()
    with open(XLSX_PATH, "wb") as f:
        f.write(data)
    print(f"  Saved to {XLSX_PATH} ({len(data)} bytes)")


def parse_xlsx():
    """Parse the Excel file and extract district math curriculum data.

    Returns list of dicts with district_name, math_k5, math_68, etc.
    """
    try:
        import openpyxl
    except ImportError:
        print("ERROR: openpyxl required. Install with: pip install openpyxl")
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, read_only=True, data_only=True)

    # List all sheet names to understand the structure
    print(f"  Sheets: {wb.sheetnames}")

    results = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Find header row (look for "District" or similar)
        header_row = None
        header_idx = None
        for i, row in enumerate(rows):
            row_str = [str(c).lower() if c else "" for c in row]
            if any("district" in c for c in row_str):
                header_row = row_str
                header_idx = i
                break

        if header_row is None:
            print(f"  Sheet '{sheet_name}': no header row found, showing first 3 rows:")
            for row in rows[:3]:
                print(f"    {row}")
            continue

        print(f"  Sheet '{sheet_name}': header at row {header_idx}, {len(rows) - header_idx - 1} data rows")
        print(f"    Columns: {[c for c in header_row if c]}")

        # Parse data rows
        for row in rows[header_idx + 1:]:
            if not row or not row[0]:
                continue

            record = {}
            for j, col_name in enumerate(header_row):
                if j < len(row) and col_name:
                    record[col_name] = row[j]

            results.append((sheet_name, record))

    wb.close()
    return results


def build_ne_district_lookup(conn):
    """Build NE district name -> LEAID lookup."""
    cur = conn.cursor()
    cur.execute("SELECT leaid, district_name FROM districts WHERE state = 'NE'")
    rows = cur.fetchall()

    lookup = {}
    for leaid, name in rows:
        lookup[name.lower().strip()] = leaid
        # Remove common suffixes
        for suffix in [" public schools", " public school", " schools", " school district"]:
            if name.lower().endswith(suffix):
                lookup[name.lower().replace(suffix, "").strip()] = leaid

    return lookup


def match_ne_district(name, lookup):
    """Match a Nebraska district name to LEAID."""
    name_lower = name.strip().lower()

    if name_lower in lookup:
        return lookup[name_lower]

    # Try adding "public schools"
    for suffix in [" public schools", " public school"]:
        if (name_lower + suffix) in lookup:
            return lookup[name_lower + suffix]

    # Substring match
    for key, leaid in lookup.items():
        if key in name_lower or name_lower in key:
            if len(key) > 3:
                return leaid

    return None


def extract_curriculum_value(val, other_val=None):
    """Clean a curriculum cell value, falling back to 'other' column if needed."""
    if not val:
        return None
    val_str = str(val).strip()
    if val_str.lower() in ("n/a", "na", "", "none", "not applicable"):
        return None
    # If it's "Other" and there's an other column value, use that
    if val_str.lower() == "other" and other_val:
        other_str = str(other_val).strip()
        if other_str.lower() not in ("n/a", "na", "", "none"):
            return other_str
    if val_str.lower() == "other":
        return None
    return val_str


def main():
    conn = sqlite3.connect(DB_PATH)
    mapping = load_normalization_map()

    # Download and parse
    download_xlsx()
    records = parse_xlsx()
    print(f"\nTotal records from Excel: {len(records)}")

    if not records:
        print("No records found. Check the Excel file structure.")
        conn.close()
        return

    # Build lookup
    ne_lookup = build_ne_district_lookup(conn)
    print(f"\nOur NE districts in DB: {len(ne_lookup)}")

    # Use the "District Math" sheet specifically — it has columns:
    # district, esu, math k-2, math k-2 other, k-2 year adopted,
    # math 3-5, math 3-5 other, 3-5 adopt year,
    # math 6-8, math 6-8 other, 6-8 adopt year,
    # math 9-12, math 9-12 other, adopt year hs2
    #
    # We can also use "District materials" which has the same math columns.
    # Let's use "District Math" as primary.

    district_curricula = {}  # district_name -> {k5_raw, g68_raw}

    for sheet_name, record in records:
        # Only process "District Math" sheet (or "District materials" as fallback)
        if sheet_name not in ("District Math", "District materials"):
            continue

        district_name = record.get("district")
        if not district_name:
            continue
        district_name = str(district_name).strip()

        # Skip if we already have this district from "District Math"
        if district_name in district_curricula and sheet_name == "District materials":
            continue

        # Extract K-5 math (combine K-2 and 3-5 — use 3-5 as primary for K-5)
        math_k2 = extract_curriculum_value(
            record.get("math k-2"),
            record.get("math k-2 other"),
        )
        math_35 = extract_curriculum_value(
            record.get("math 3-5"),
            record.get("math 3-5 other"),
        )
        # Use 3-5 as primary K-5 curriculum, fall back to K-2
        k5_raw = math_35 or math_k2

        # Extract 6-8 math
        g68_raw = extract_curriculum_value(
            record.get("math 6-8"),
            record.get("math 6-8 other"),
        )

        if k5_raw or g68_raw:
            district_curricula[district_name] = {"k5_raw": k5_raw, "g68_raw": g68_raw}

    print(f"Districts with math curriculum: {len(district_curricula)}")

    # Show a few examples
    for name, data in list(district_curricula.items())[:3]:
        print(f"  {name}: K-5='{data['k5_raw']}', 6-8='{data['g68_raw']}'")

    # Match and insert
    matched = 0
    inserted = 0
    unmatched = []
    cur = conn.cursor()

    for district_name, curricula in sorted(district_curricula.items()):
        leaid = match_ne_district(district_name, ne_lookup)
        if not leaid:
            unmatched.append(district_name)
            continue

        matched += 1

        k5_raw = curricula["k5_raw"]
        g68_raw = curricula["g68_raw"]

        k5_norm = normalize_curriculum_name(k5_raw, mapping)[0] if k5_raw else None
        g68_norm = normalize_curriculum_name(g68_raw, mapping)[0] if g68_raw else None

        if not k5_raw and not g68_raw:
            continue

        cur.execute("""
            INSERT OR REPLACE INTO curriculum
            (leaid, k5_curriculum, k5_normalized, grade68_curriculum, grade68_normalized,
             source_tier, source_url, confidence, date_collected)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            leaid, k5_raw, k5_norm, g68_raw, g68_norm,
            SOURCE_TIER, SOURCE_URL, 1.0, TODAY,
        ))
        inserted += 1

    conn.commit()

    print(f"\nResults:")
    print(f"  Matched: {matched}")
    print(f"  Inserted: {inserted}")
    print(f"  Unmatched: {len(unmatched)}")
    if unmatched[:10]:
        print(f"  Sample unmatched: {unmatched[:10]}")

    # Summary
    print(f"\nNE Curriculum distribution (K-5):")
    cur.execute("""
        SELECT k5_normalized, COUNT(*) as cnt
        FROM curriculum WHERE source_tier = 1 AND k5_normalized IS NOT NULL
        AND leaid IN (SELECT leaid FROM districts WHERE state = 'NE')
        GROUP BY k5_normalized ORDER BY cnt DESC LIMIT 15
    """)
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]} districts")

    print(f"\nNE Curriculum distribution (6-8):")
    cur.execute("""
        SELECT grade68_normalized, COUNT(*) as cnt
        FROM curriculum WHERE source_tier = 1 AND grade68_normalized IS NOT NULL
        AND leaid IN (SELECT leaid FROM districts WHERE state = 'NE')
        GROUP BY grade68_normalized ORDER BY cnt DESC LIMIT 15
    """)
    for row in cur.fetchall():
        print(f"  {row[0]}: {row[1]} districts")

    conn.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
